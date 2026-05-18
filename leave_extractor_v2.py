import streamlit as st
import google.generativeai as genai
from PIL import Image
import json
import pandas as pd
import io
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ==========================================
# 1. CONFIGURATION
# ==========================================
API_KEY = st.secrets["Gen_API"]["API_KEY"] # Replace with your actual key
genai.configure(api_key=API_KEY)

# Initialize both models based on your previous code
model_flash = genai.GenerativeModel('gemini-3-flash-preview')
model_pro = genai.GenerativeModel('gemini-1.5-pro')

st.set_page_config(page_title='Zenxin Document Extractor', layout='wide')

# ==========================================
# 2. SIDEBAR - MODE SELECTION
# ==========================================
st.sidebar.title("Settings")
st.sidebar.write("Select Mode")
app_mode = st.sidebar.radio(
    label="Select Mode",
    label_visibility="collapsed", # Hides the duplicate label to match your image
    options=("Attendance Report OCR", "Leave/Overtime OCR")
)

# ==========================================
# 3. MAIN UI & IMAGE INPUT
# ==========================================
st.title(f'📄 {app_mode}')
st.write('Upload or capture a photo of the document.')

upload_option = st.radio('Input Method:', ('Camera', 'Upload Image'), horizontal=True)

# ---> THE FIX: Initialize it here so it always exists <---
uploaded_files = [] 

if upload_option == 'Camera':
    # 1. Initialize session state to keep track of whether the camera should be ON or OFF
    if 'camera_active' not in st.session_state:
        st.session_state.camera_active = False

    # 2. Show the "Open Camera" button if the camera is currently OFF
    if not st.session_state.camera_active:
        if st.button('📸 Open Camera', use_container_width=True):
            st.session_state.camera_active = True
            st.rerun() # Refresh the page to load the camera
    
    # 3. Show the camera ONLY if active
    else:
        if st.button('❌ Close Camera', use_container_width=True):
            st.session_state.camera_active = False
            st.rerun() # Refresh to hide the camera
            
        camera_file = st.camera_input('Capture Report', label_visibility="collapsed")
        # Only overwrite our empty list if a photo was actually taken
        if camera_file:
            uploaded_files = [camera_file]

else:
    # Ensure camera turns off if they switch to "Upload Image"
    st.session_state.camera_active = False 
    
    # File uploader configured to accept multiple images
    uploader_files = st.file_uploader(
        'Upload Report Pages', 
        type=['jpg', 'png', 'jpeg'], 
        accept_multiple_files=True
    )
    # Only overwrite our empty list if files were uploaded
    if uploader_files:
        uploaded_files = uploader_files

# ==========================================
# 4. ANALYSIS & DISPLAY
# ==========================================
if uploaded_files:
    # Convert uploaded files to PIL Images
    images = [Image.open(file) for file in uploaded_files]
    
    col1, col2 = st.columns([1, 1.5])
    
    with col1:
        st.image(images, use_container_width=True)
    
    with col2:
        # ------------------------------------------
        # MODE A: ATTENDANCE REPORT
        # ------------------------------------------
        if app_mode == "Attendance Report OCR":
            if st.button('Extract Data & Generate Excel', use_container_width=True):
                with st.spinner(f'Batch processing {len(images)} images in a SINGLE request...'):
                    
                    # 1. Prompt updated to handle MULTIPLE images and return a JSON ARRAY
                    prompt = f'''
                    You are an expert at transcribing messy handwritten notes on printed timesheets. 
                    I am providing you with {len(images)} images of Attendance Reports.
                    You MUST extract the data for EVERY SINGLE IMAGE provided.
                    
                    Return ONLY a valid JSON ARRAY containing exactly {len(images)} objects. 
                    Object 1 corresponds to IMAGE 1, Object 2 corresponds to IMAGE 2, and so on.
                    Do NOT include markdown formatting like ```json in the output.

                    ### 🚨 HANDWRITING DICTIONARY & RULES 🚨
                    1. Known Codes: 'UPL', 'UPH', 'MC', 'AL', 'OT', 'OFF'.
                    2. Column Locations: Handwriting is usually over 'Resume', 'Out', and 'OT'.
                    3. Circled Numbers: Transcribe as "UPL - 1 (Circled [Number])".
                    4. Crossed-out Text: If a number is crossed out, include it like "~~5.5 Hours~~".
                    5. Checkmarks: Note any checkmarks (✓).
                    6. Margin Notes: Look for Malay text like "Lambat masuk" or math.

                    Return EXACTLY this JSON structure (as an Array of objects):
                    [
                        {{
                          "employee": {{
                            "User ID": "String", "Name": "String", "Department": "String", "Date Range": "String"
                          }},
                          "timesheet": [
                            {{"Date": "String", "Day": "String", "In": "String", "Out": "String", "Work Hrs": "String", "Remarks": "String"}}
                          ],
                          "summary": [
                            "String"
                          ],
                          "notes": [
                            "String"
                          ]
                        }}
                    ]
                    '''
                    
                    try:
                        # 2. Interleave Text and Images
                        request_content = [prompt]
                        for i, img in enumerate(images):
                            request_content.append(f"--- START OF IMAGE {i + 1} ---")
                            request_content.append(img)
                            request_content.append(f"--- END OF IMAGE {i + 1} ---")

                        # 3. Make ONE single API call
                        response = model_flash.generate_content(
                            request_content, 
                            generation_config=genai.types.GenerationConfig(temperature=0.0) # 0.0 for strict format adherence
                        )
                        
                        clean_json = response.text.replace('```json', '').replace('```', '').strip()
                        extracted_data_list = json.loads(clean_json)

                        # Safeguard: Ensure it returned a list
                        if not isinstance(extracted_data_list, list):
                            extracted_data_list = [extracted_data_list]

                        st.success(f'Extraction Complete! Processed {len(extracted_data_list)} records in 1 request.')
                        
                        # --- PREPARE MASTER LISTS FOR EXCEL ---
                        all_timesheets_df = []
                        all_employees_df = []
                        
                        # --- BUILD THE UI ---
                        for idx, sheet in enumerate(extracted_data_list):
                            emp_name = sheet.get('employee', {}).get('Name', 'Unknown')
                            
                            with st.expander(f"📄 Record {idx + 1}: {emp_name}", expanded=True):
                                st.markdown("### Employee Information")
                                for key, value in sheet.get('employee', {}).items():
                                    st.markdown(f"* **{key}:** {value}")
                                    
                                st.markdown("### Timesheet Data")
                                df_timesheet = pd.DataFrame(sheet.get('timesheet', []))
                                st.dataframe(df_timesheet, use_container_width=True)
                                
                                # Display Summary and Notes safely
                                summary_list = sheet.get('summary', [])
                                notes_list = sheet.get('notes', [])
                                
                                col_sum1, col_sum2 = st.columns(2)
                                with col_sum1:
                                    if summary_list:
                                        st.markdown("### Summary")
                                        for s in summary_list:
                                            st.markdown(f"* {s}")
                                with col_sum2:
                                    if notes_list:
                                        st.markdown("### Notes")
                                        for n in notes_list:
                                            st.markdown(f"* {n}")

                            # --- COMPILE DATA FOR MASTER EXCEL ---
                            if not df_timesheet.empty:
                                # Inject employee info into the timesheet rows
                                df_timesheet.insert(0, 'Date Range', sheet.get('employee', {}).get('Date Range', 'Unknown'))
                                df_timesheet.insert(0, 'Department', sheet.get('employee', {}).get('Department', 'Unknown'))
                                df_timesheet.insert(0, 'User ID', sheet.get('employee', {}).get('User ID', 'Unknown'))
                                df_timesheet.insert(0, 'Name', emp_name)
                                all_timesheets_df.append(df_timesheet)
                            
                            # Create a flat row for the employee info sheet, merging the arrays into strings
                            emp_row = sheet.get('employee', {}).copy()
                            emp_row['Summary'] = " | ".join(summary_list) if isinstance(summary_list, list) else str(summary_list)
                            emp_row['Notes'] = " | ".join(notes_list) if isinstance(notes_list, list) else str(notes_list)
                            all_employees_df.append(emp_row)

                        # --- GENERATE MASTER EXCEL FILE ---
                        if all_employees_df:
                            final_records = pd.concat(all_employees_df, ignore_index=True)
                            final_summaries = pd.DataFrame(all_employees_df)

                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                final_records.to_excel(writer, index=False, sheet_name='All OT Records')
                                final_summaries.to_excel(writer, index=False, sheet_name='All Summaries')
                                
                                # --- NEW: APPLY EXCEL TABLE FORMATTING & LINES ---
                                workbook = writer.book
                                
                                for sheet_name in ['All OT Records', 'All Summaries']:
                                    worksheet = writer.sheets[sheet_name]
                                    max_row = worksheet.max_row
                                    max_col = worksheet.max_column
                                    
                                    # 1. Define the table range (e.g., "A1:G15")
                                    table_range = f"A1:{get_column_letter(max_col)}{max_row}"
                                    
                                    # 2. Create the Table object
                                    safe_name = sheet_name.replace(" ", "_") # Excel table names can't have spaces
                                    tab = Table(displayName=safe_name, ref=table_range)
                                    
                                    # 3. Apply a default Excel style (Adds lines, banded rows, and header filters)
                                    style = TableStyleInfo(
                                        name="TableStyleMedium9", 
                                        showRowStripes=True,
                                        showColumnStripes=False
                                    )
                                    tab.tableStyleInfo = style
                                    worksheet.add_table(tab)
                                    
                                    # 4. Auto-adjust column widths so text isn't hidden
                                    for col in worksheet.columns:
                                        max_length = 0
                                        column = col[0].column_letter # Get the column name
                                        for cell in col:
                                            try:
                                                if len(str(cell.value)) > max_length:
                                                    max_length = len(str(cell.value))
                                            except:
                                                pass
                                        adjusted_width = (max_length + 2)
                                        worksheet.column_dimensions[column].width = adjusted_width

                            excel_data = output.getvalue()
                            
                            st.download_button(
                                label="📥 Download Master Excel (.xlsx)",
                                data=excel_data,
                                file_name="Master_Leave_OT_Report.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                use_container_width=True
                            )
                            
                    except Exception as e:
                        st.error(f'Error processing batch. Please try again.')
                        st.expander("View Error Log").write(e)
        # ------------------------------------------
        # MODE B: LEAVE/OVERTIME RECORD
        # ------------------------------------------
        # ------------------------------------------
        # MODE B: LEAVE/OVERTIME RECORD
        # ------------------------------------------
        elif app_mode == "Leave/Overtime OCR":
            if st.button('Extract Overtime Data & Generate Excel', use_container_width=True):
                with st.spinner(f'Batch processing {len(images)} images in a SINGLE request...'):
                    
                    prompt = f"""
                    You are a data extraction assistant. I am providing you with {len(images)} images of Leave/Overtime Records.
                    You MUST extract the data for EVERY SINGLE IMAGE provided.
                    
                    Return ONLY a valid JSON ARRAY containing exactly {len(images)} objects. 
                    Object 1 corresponds to IMAGE 1, Object 2 corresponds to IMAGE 2, and so on.
                    Do NOT include markdown formatting like ```json.

                    ### 🚨 EXTRACTION RULES 🚨
                    1. Extract all rows from the table. Empty cells should be an empty string "".
                    2. Convert fractions in hours to decimals (e.g., "1 ½" becomes 1.5).

                    Return EXACTLY this JSON structure (as an Array of objects):
                    [
                        {{
                          "employee": {{
                            "Company": "String", "Name": "String", "Position": "String", "Month": "String", "Emp No": "String"
                          }},
                          "records": [
                            {{"Date": "String", "Type (Leave/OT)": "String", "Reason/Remark": "String", "Time From": "String", "Time To": "String", "Hours": "Number"}}
                          ],
                          "signatures": {{
                            "Confirmed By": "String", "Approval By": "String"
                          }}
                        }}
                    ]
                    """

                    try:
                        request_content = [prompt]
                        for i, img in enumerate(images):
                            request_content.append(f"--- START OF IMAGE {i + 1} ---")
                            request_content.append(img)
                            request_content.append(f"--- END OF IMAGE {i + 1} ---")

                        response = model_flash.generate_content(
                            request_content,
                            generation_config=genai.types.GenerationConfig(temperature=0.0) 
                        )
                        
                        clean_json = response.text.replace('```json', '').replace('```', '').strip()
                        extracted_data_list = json.loads(clean_json)

                        if not isinstance(extracted_data_list, list):
                            extracted_data_list = [extracted_data_list]

                        # ==========================================
                        # --- NEW: GROUP PAGES BY EMPLOYEE NAME ---
                        # ==========================================
                        consolidated_data = {}
                        for sheet in extracted_data_list:
                            emp_info = sheet.get('employee', {})
                            # Normalize name to Title Case so "Akmal Hakim" matches "AKMAL HAKIM"
                            emp_name = str(emp_info.get('Name', 'Unknown')).strip().title()
                            
                            if emp_name not in consolidated_data:
                                # First time seeing this employee, create their profile
                                consolidated_data[emp_name] = {
                                    'employee': emp_info,
                                    'records': [],
                                    'signatures': sheet.get('signatures', {})
                                }
                            else:
                                # We found another page for this same employee!
                                # Let's update the month label so it says "MAR 2026 & APRIL 2026"
                                existing_month = str(consolidated_data[emp_name]['employee'].get('Month', ''))
                                new_month = str(emp_info.get('Month', ''))
                                if new_month and new_month not in existing_month:
                                    consolidated_data[emp_name]['employee']['Month'] = f"{existing_month} & {new_month}"

                            # Add this page's rows to the employee's master list of rows
                            consolidated_data[emp_name]['records'].extend(sheet.get('records', []))

                        st.success(f'Extraction Complete! Processed {len(images)} pages into {len(consolidated_data)} employee records.')
                        
                        all_records_df = []
                        all_summaries_df = []

                        # --- BUILD THE UI & CALCULATE MATH IN PYTHON ---
                        # Notice we now loop through consolidated_data instead of extracted_data_list
                        for idx, (emp_name, sheet) in enumerate(consolidated_data.items()):
                            with st.expander(f"📄 Record {idx + 1}: {emp_name}", expanded=True):
                                
                                st.markdown("### Employee Information")
                                for key, value in sheet.get('employee', {}).items():
                                    st.markdown(f"* **{key}:** {value}")
                                    
                                st.markdown("### Overtime & Leave Records")
                                df_records = pd.DataFrame(sheet.get('records', []))
                                
                                # --- EXACT PANDAS CALCULATIONS ---
                                calculated_summary = {
                                    "Total OT Hours": 0.0,
                                    "Total OT Days": 0,
                                    "Total AL Days": 0,
                                    "Total UPL Days": 0,
                                    "Total MC Days": 0
                                }

                                if not df_records.empty:
                                    
                                    def calc_hours(row):
                                        start_str = str(row.get('Time From', '')).strip()
                                        end_str = str(row.get('Time To', '')).strip()
                                        
                                        def to_decimal(t):
                                            t = t.replace('.', ':').replace(',', ':').lower().replace('am', '').replace('pm', '').strip()
                                            if not t: return None
                                            if ':' not in t:
                                                try: return float(t)
                                                except: return None
                                            parts = t.split(':')
                                            try: return float(parts[0]) + (float(parts[1]) / 60.0)
                                            except: return None
                                            
                                        start_val = to_decimal(start_str)
                                        end_val = to_decimal(end_str)
                                        
                                        if start_val is not None and end_val is not None:
                                            diff = end_val - start_val
                                            if diff < 0:
                                                diff += 12 
                                            return round(diff, 2)
                                        
                                        try:
                                            return float(row.get('Hours', 0))
                                        except:
                                            return 0.0

                                    df_records['Calculated Hours'] = df_records.apply(calc_hours, axis=1)
                                    
                                    if 'Type (Leave/OT)' in df_records.columns:
                                        type_col = df_records['Type (Leave/OT)'].fillna('').astype(str).str.upper()
                                        is_ot_condition = type_col.str.contains('OT') | ((type_col == '') & (df_records['Calculated Hours'] > 0))
                                        
                                        calculated_summary["Total OT Hours"] = df_records[is_ot_condition]['Calculated Hours'].sum()
                                        calculated_summary["Total OT Days"] = is_ot_condition.sum()
                                        
                                        calculated_summary["Total AL Days"] = type_col.str.contains('AL').sum()
                                        calculated_summary["Total UPL Days"] = type_col.str.contains('UPL').sum()
                                        calculated_summary["Total MC Days"] = type_col.str.contains('MC').sum()

                                st.dataframe(df_records, use_container_width=True)
                                
                                col_sum1, col_sum2 = st.columns(2)
                                with col_sum1:
                                    st.markdown("### Leave/OT Summary (Calculated)")
                                    for key, value in calculated_summary.items():
                                        st.markdown(f"* **{key}:** {value}")
                                with col_sum2:
                                    st.markdown("### Signatures")
                                    for key, value in sheet.get('signatures', {}).items():
                                        st.markdown(f"* **{key}:** {value}")

                            # --- COMPILE DATA FOR MASTER EXCEL ---
                            emp_no = sheet.get('employee', {}).get('Emp No', 'Unknown')
                            month = sheet.get('employee', {}).get('Month', 'Unknown')

                            if not df_records.empty:
                                df_records.insert(0, 'Month', month)
                                df_records.insert(0, 'Emp No', emp_no)
                                df_records.insert(0, 'Employee Name', emp_name)
                                all_records_df.append(df_records)

                            summary_row = sheet.get('employee', {}).copy()
                            summary_row.update(calculated_summary) 
                            all_summaries_df.append(summary_row)

                        # --- GENERATE MASTER EXCEL FILE ---
                        if all_records_df:
                            final_records = pd.concat(all_records_df, ignore_index=True)
                            final_summaries = pd.DataFrame(all_summaries_df)

                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                final_records.to_excel(writer, index=False, sheet_name='All OT Records')
                                final_summaries.to_excel(writer, index=False, sheet_name='All Summaries')
                                
                                workbook = writer.book
                                
                                for sheet_name in ['All OT Records', 'All Summaries']:
                                    worksheet = writer.sheets[sheet_name]
                                    max_row = worksheet.max_row
                                    max_col = worksheet.max_column
                                    
                                    table_range = f"A1:{get_column_letter(max_col)}{max_row}"
                                    safe_name = sheet_name.replace(" ", "_") 
                                    tab = Table(displayName=safe_name, ref=table_range)
                                    
                                    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
                                    tab.tableStyleInfo = style
                                    worksheet.add_table(tab)
                                    
                                    for col in worksheet.columns:
                                        max_length = 0
                                        column = col[0].column_letter 
                                        for cell in col:
                                            try:
                                                if len(str(cell.value)) > max_length:
                                                    max_length = len(str(cell.value))
                                            except:
                                                pass
                                        adjusted_width = (max_length + 2)
                                        worksheet.column_dimensions[column].width = adjusted_width

                            excel_data = output.getvalue()
                            
                            st.download_button(
                                label="📥 Download Master Excel (.xlsx)",
                                data=excel_data,
                                file_name="Master_Leave_OT_Report.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                use_container_width=True
                            )
                            
                    except Exception as e:
                        st.error(f"Error processing batch: {e}")
                        st.expander("View Error Log").write(e)
